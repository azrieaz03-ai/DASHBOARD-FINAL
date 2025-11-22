from flask import Flask, render_template, request, redirect, url_for, session, Response, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from datetime import datetime, date
from openpyxl import Workbook
from collections import defaultdict
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io
import os
from zoneinfo import ZoneInfo

# Fungsi waktu WIB yang lebih robust
def waktu_wib():
    return datetime.now(ZoneInfo("Asia/Jakarta"))

app = Flask(__name__)
app.secret_key = os.urandom(32)

# --- konfigurasi database ---
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:hendri@localhost/bakery_new'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ------------------ USER ------------------
class User(db.Model):
    __tablename__ = 'user'
    idUser  = db.Column(db.Integer, primary_key=True)
    Username = db.Column(db.String(45), unique=True, nullable=False)
    Password = db.Column(db.String(45), nullable=False)
    Role = db.Column(db.Enum('Owner','Kasir'), nullable=False)
    Nama = db.Column(db.String(45), nullable=False)

    produksi = db.relationship('ProduksiHarian', backref='user', lazy=True)
    transaksi = db.relationship('TransaksiPenjualan', backref='user', lazy=True)
    ringkasan = db.relationship('RingkasanBaru', backref='user', lazy=True)

# ------------------ JENIS ROTI ------------------
class JenisRoti(db.Model):
    __tablename__ = 'roti'
    idRoti = db.Column(db.Integer, primary_key=True)
    Nama_Roti = db.Column(db.String(45), unique=True, nullable=False)
    Harga = db.Column(db.Numeric(10), nullable=False)

    produksi = db.relationship('ProduksiHarian', backref='roti', lazy=True)
    detail_transaksi = db.relationship('DetailTransaksi', backref='roti', lazy=True)

# ------------------ PRODUKSI ------------------
class ProduksiHarian(db.Model):
    __tablename__ = 'produksi'
    idProduksi = db.Column(db.Integer, primary_key=True)
    idRoti = db.Column(db.Integer, db.ForeignKey('roti.idRoti'), nullable=False)
    idUser = db.Column(db.Integer, db.ForeignKey('user.idUser'), nullable=False)
    Jumlah_Produksi = db.Column(db.Integer, nullable=False)
    Tanggal_Produksi = db.Column(db.DateTime, default=waktu_wib)

    ringkasan = db.relationship('RingkasanBaru', backref='produksi', lazy=True)

# ------------------ TRANSAKSI ------------------
class TransaksiPenjualan(db.Model):
    __tablename__ = 'transaksi_penjualan'
    idTransaksi_Penjualan = db.Column(db.Integer, primary_key=True)
    idUser = db.Column(db.Integer, db.ForeignKey('user.idUser'), nullable=False)
    Tanggal_Transaksi = db.Column(db.DateTime, default=waktu_wib)
    Total_Harga = db.Column(db.Numeric(10), nullable=False)
    Uang_Diterima = db.Column(db.Numeric(10), nullable=False)
    Kembalian = db.Column(db.Numeric(10), nullable=False)

    detail_transaksi = db.relationship('DetailTransaksi', backref='transaksi', lazy=True)
    ringkasan = db.relationship('RingkasanBaru', backref='transaksi', lazy=True)

# ------------------ DETAIL TRANSAKSI ------------------
class DetailTransaksi(db.Model):
    __tablename__ = 'detail_transaksi'
    idDetail_Transaksi = db.Column(db.Integer, primary_key=True)
    id_transaksi = db.Column(db.Integer, db.ForeignKey('transaksi_penjualan.idTransaksi_Penjualan'), nullable=False)
    id_roti = db.Column(db.Integer, db.ForeignKey('roti.idRoti'), nullable=False)
    Jumlah = db.Column(db.Integer, nullable=False)
    SubTotal = db.Column(db.Numeric(10), nullable=False)

# ------------------ RINGKASAN BARU ------------------
class RingkasanBaru(db.Model):
    __tablename__ = 'ringkasan_baru'
    idRingkasan = db.Column(db.Integer, primary_key=True)
    idUser = db.Column(db.Integer, db.ForeignKey('user.idUser'), nullable=False)
    idProduksi = db.Column(db.Integer, db.ForeignKey('produksi.idProduksi'), nullable=False)
    idTransaksi_Penjualan = db.Column(db.Integer, db.ForeignKey('transaksi_penjualan.idTransaksi_Penjualan'), nullable=True)
    Tanggal = db.Column(db.DateTime, default=waktu_wib)
    Total_Produksi = db.Column(db.Integer, nullable=False)
    Total_Terjual = db.Column(db.Integer, nullable=False)
    Stok_Aktual = db.Column(db.Integer, nullable=False)
    Total_Uang_Masuk = db.Column(db.Numeric(10), nullable=False)

# --- Membuat semua tabel ---
with app.app_context():
    db.create_all()

def update_ringkasan_terbaru(idRoti):
    p = ProduksiHarian.query.filter_by(idRoti=idRoti)\
        .order_by(ProduksiHarian.Tanggal_Produksi.desc()).first()
    
    # Ambil stok terakhir dari ringkasan sebelumnya
    rk_last = RingkasanBaru.query.join(ProduksiHarian)\
        .filter(ProduksiHarian.idRoti==idRoti)\
        .order_by(RingkasanBaru.Tanggal.desc(), RingkasanBaru.idRingkasan.desc())\
        .first()
    
    stok_sebelumnya = rk_last.Stok_Aktual if rk_last else 0
    total_terjual = 0  # bisa diubah kalau mau hitung penjualan sampai saat ini
    total_uang_masuk = 0
    
    rk = RingkasanBaru(
        idUser=p.idUser,
        idProduksi=p.idProduksi,
        idTransaksi_Penjualan=None,
        Tanggal=p.Tanggal_Produksi,
        Total_Produksi=p.Jumlah_Produksi,
        Total_Terjual=total_terjual,
        Stok_Aktual=stok_sebelumnya + p.Jumlah_Produksi - total_terjual,
        Total_Uang_Masuk=total_uang_masuk
    )
    db.session.add(rk)
    db.session.commit()

# ------------------ ROUTES ------------------
@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(Username=username, Password=password).first()
        if user:
            session['user_id'] = user.idUser
            session['role'] = user.Role
            session['username'] = user.Username
            if user.Role == 'Owner':
                return redirect(url_for('dashboard_owner'))
            elif user.Role == 'Kasir':
                return redirect(url_for('dashboard_kasir'))
        else:
            error = "Username atau password salah"
    return render_template("login/login.html", error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/esp32/input_produksi', methods=['POST'])
def input_produksi_esp32():
    data = request.get_json()
    if not data:
        return jsonify({"message": "Data JSON tidak ditemukan"}), 400

    idRoti = data.get('idRoti')
    idUser = data.get('idUser')
    Jumlah_Produksi = data.get('Jumlah_Produksi')
    tanggal_input = data.get('Tanggal_Produksi')

    if not idRoti or not idUser or not Jumlah_Produksi:
        return jsonify({"message": "Field idRoti, idUser, dan Jumlah_Produksi wajib diisi"}), 400

    if tanggal_input:
        try:
            Tanggal_Produksi = datetime.fromisoformat(tanggal_input)
        except ValueError:
            return jsonify({"message": "Format Tanggal_Produksi harus YYYY-MM-DDTHH:MM:SS"}), 400
    else:
        Tanggal_Produksi = waktu_wib()

    produksi_baru = ProduksiHarian(
        idRoti=idRoti,
        idUser=idUser,
        Jumlah_Produksi=Jumlah_Produksi,
        Tanggal_Produksi=Tanggal_Produksi
    )
    db.session.add(produksi_baru)
    db.session.commit()

    update_ringkasan_terbaru(idRoti)

    return jsonify({"message": "Data produksi berhasil disimpan!"}), 200

@app.route('/dashboard_owner')
def dashboard_owner():
    if 'role' not in session or session['role'] != 'Owner':
        return redirect(url_for('login'))

    user = session.get('username')
    return render_template("owner/owner.html", user=user)

@app.route('/data_produksi')
def data_produksi():
    if 'role' not in session or session['role'] != 'Owner':
        return redirect(url_for('login'))

    selected_date = request.args.get('tanggal', None)
    periode = request.args.get('periode', 'none')
    download = request.args.get('download', None)
    data = []

    if selected_date:
        if periode == 'none':
            # Data hari ini, stok = stok kemarin + produksi – terjual
            query = text("""
                SELECT
                    r.Nama_Roti,
                    SUM(rk.Total_Produksi) AS Total_Produksi,
                    SUM(rk.Total_Terjual) AS Total_Terjual,
                    COALESCE((
                        SELECT rk2.Stok_Aktual
                        FROM ringkasan_baru rk2
                        JOIN produksi p2 ON rk2.idProduksi = p2.idProduksi
                        WHERE p2.idRoti = r.idRoti
                        AND DATE(rk2.Tanggal) < :tanggal
                        ORDER BY rk2.Tanggal DESC, rk2.idRingkasan DESC
                        LIMIT 1
                    ),0)
                    + SUM(rk.Total_Produksi)
                    - SUM(rk.Total_Terjual) AS Stok_Aktual,
                    SUM(rk.Total_Uang_Masuk) AS Total_Uang_Masuk
                FROM ringkasan_baru rk
                JOIN produksi p ON rk.idProduksi = p.idProduksi
                JOIN roti r ON p.idRoti = r.idRoti
                WHERE DATE(rk.Tanggal) = :tanggal
                GROUP BY r.idRoti
                ORDER BY r.idRoti
            """)
            data = db.session.execute(query, {"tanggal": selected_date}).fetchall()
        else:
            hari = int(periode)
            # Data kumulatif periode tertentu
            query = text("""
                SELECT
                    r.Nama_Roti,
                    SUM(rk.Total_Produksi) AS Total_Produksi,
                    SUM(rk.Total_Terjual) AS Total_Terjual,
                    COALESCE((
                        SELECT rk2.Stok_Aktual
                        FROM ringkasan_baru rk2
                        JOIN produksi p2 ON rk2.idProduksi = p2.idProduksi
                        WHERE p2.idRoti = r.idRoti
                        AND DATE(rk2.Tanggal) < DATE_SUB(:tanggal, INTERVAL :hari DAY)
                        ORDER BY rk2.Tanggal DESC, rk2.idRingkasan DESC
                        LIMIT 1
                    ),0)
                    + SUM(rk.Total_Produksi)
                    - SUM(rk.Total_Terjual) AS Stok_Aktual,
                    SUM(rk.Total_Uang_Masuk) AS Total_Uang_Masuk
                FROM ringkasan_baru rk
                JOIN produksi p ON rk.idProduksi = p.idProduksi
                JOIN roti r ON p.idRoti = r.idRoti
                WHERE DATE(rk.Tanggal) BETWEEN DATE_SUB(:tanggal, INTERVAL :hari DAY) AND :tanggal
                GROUP BY r.idRoti
                ORDER BY r.idRoti
            """)
            data = db.session.execute(query, {"tanggal": selected_date, "hari": hari}).fetchall()
    # Jika request download Excel
    if download == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Produksi"

        # Keterangan tanggal & periode
        periode_text = {
            'none': 'Tidak Kumulatif',
            '7': 'Kumulatif 7 Hari',
            '30': 'Kumulatif 30 Hari',
            '90': 'Kumulatif 90 Hari'
        }
        ws.append([f"Tanggal: {selected_date}"])
        ws.append([f"Periode: {periode_text.get(periode, 'Unknown')}"])
        ws.append([])  # baris kosong

        # Header tabel
        headers = ['Jenis Roti', 'Produksi', 'Terjual', 'Stok', 'Uang Masuk']
        ws.append(headers)

        # Styling untuk header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Kuning
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[4]:  # Baris 4 adalah header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
            cell.border = thin_border

        # Data tabel
        for row in data:
            ws.append([row.Nama_Roti, row.Total_Produksi, row.Total_Terjual, row.Stok_Aktual, row.Total_Uang_Masuk])

        # Tambahkan border ke semua sel tabel (mulai dari baris 4 sampai akhir)
        for row_cells in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row_cells:
                cell.border = thin_border

        # Auto-fit column width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Simpan ke buffer dan kirim ke browser
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            download_name=f"data_produksi_{selected_date}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # Render HTML biasa
    return render_template(
        "owner/data produksi.html",
        data=data,
        selected_date=selected_date,
        periode=periode
    )

# ROUTE HALAMAN LAPORAN PENJUALAN
@app.route('/laporan_penjualan')
def laporan_penjualan():
    if 'role' not in session or session['role'] != 'Owner':
        return redirect(url_for('login'))
    return render_template("owner/laporan penjualan.html")

@app.route('/api/data_laporan')
def api_data_laporan():
    if 'role' not in session or session['role'] != 'Owner':
        return jsonify({"error": "Unauthorized"}), 401

    selected_date = request.args.get('tanggal', None)
    periode = request.args.get('periode', 'none')  # 'none', '7', '30', '90'

    labels = []
    values_terjual = []
    values_uang = []

    if not selected_date:
        return jsonify({"labels_chart": [], "values_terjual": [], "values_uang": []})

    # Ambil semua data
    if periode == 'none':
        query = text("""
            SELECT r.Nama_Roti, DATE(rk.Tanggal) AS tgl,
                   rk.Total_Terjual, rk.Total_Uang_Masuk
            FROM ringkasan_baru rk
            JOIN produksi p ON rk.idProduksi = p.idProduksi
            JOIN roti r ON p.idRoti = r.idRoti
            WHERE DATE(rk.Tanggal) = :tanggal
            ORDER BY r.idRoti
        """)
        rows = db.session.execute(query, {"tanggal": selected_date}).fetchall()
    else:
        hari = int(periode)
        query = text("""
            SELECT r.Nama_Roti, DATE(rk.Tanggal) AS tgl,
                   rk.Total_Terjual, rk.Total_Uang_Masuk
            FROM ringkasan_baru rk
            JOIN produksi p ON rk.idProduksi = p.idProduksi
            JOIN roti r ON p.idRoti = r.idRoti
            WHERE rk.Tanggal BETWEEN DATE_SUB(:tanggal, INTERVAL :hari DAY) AND :tanggal
            ORDER BY r.idRoti, DATE(rk.Tanggal)
        """)
        rows = db.session.execute(query, {"tanggal": selected_date, "hari": hari}).fetchall()

    # Grouping & sum per tanggal per roti
    grouped_terjual = defaultdict(lambda: defaultdict(int))
    grouped_uang = defaultdict(lambda: defaultdict(float))

    for nama, tgl, terjual, uang in rows:
        grouped_terjual[nama][str(tgl)] += terjual
        grouped_uang[nama][str(tgl)] += float(uang)

    for nama in grouped_terjual.keys():
        labels.append(nama)
        values_terjual.append([{"tanggal": t, "jumlah": j} for t, j in grouped_terjual[nama].items()])
        values_uang.append([{"tanggal": t, "uang": u} for t, u in grouped_uang[nama].items()])

    return jsonify({
        "labels_chart": labels,
        "values_terjual": values_terjual,
        "values_uang": values_uang
    })

# Dashboard Kasir
@app.route('/dashboard_kasir')
def dashboard_kasir():
    if 'role' not in session or session['role'] != 'Kasir':
        return redirect(url_for('login'))
    user = session.get('username')
    return render_template("kasir/kasir.html", user=user)


@app.route('/data_produksi_kasir')
def data_produksi_kasir():
    if 'role' not in session or session['role'] != 'Kasir':
        return redirect(url_for('login'))
    user = session.get('username')

    today = date.today().strftime("%Y-%m-%d")

    query = text("""
        SELECT r.Nama_Roti,
               SUM(rk.Total_Produksi) AS Total_Produksi,
               SUM(rk.Total_Terjual) AS Total_Terjual,
               COALESCE((
                    SELECT rk2.Stok_Aktual
                    FROM ringkasan_baru rk2
                    JOIN produksi p2 ON rk2.idProduksi = p2.idProduksi
                    WHERE p2.idRoti = r.idRoti
                      AND rk2.Tanggal < :tanggal
                    ORDER BY rk2.Tanggal DESC, rk2.idRingkasan DESC
                    LIMIT 1
               ),0) + SUM(rk.Total_Produksi) - SUM(rk.Total_Terjual) AS Stok_Aktual,
               SUM(rk.Total_Uang_Masuk) AS Total_Uang
        FROM ringkasan_baru rk
        JOIN produksi p ON rk.idProduksi = p.idProduksi
        JOIN roti r ON p.idRoti = r.idRoti
        WHERE DATE(rk.Tanggal) = :tanggal
        GROUP BY r.idRoti
        ORDER BY r.idRoti
    """)

    rows = db.session.execute(query, {"tanggal": today}).fetchall()

    # ✅ Ubah Row → dict agar bisa tojson
    data = []
    for row in rows:
        data.append({
            "Nama": row[0],
            "Total_Produksi": int(row[1] or 0),
            "Total_Terjual": int(row[2] or 0),
            "Stok_Aktual": int(row[3] or 0),
            "Total_Uang": int(row[4] or 0),
        })

    # ✅ Kirim "data" bukan "result"
    return render_template("kasir/data produksi kasir.html", user=user, data=data)

@app.route('/Hitung_Total')
def Hitung_Total():
    if 'role' not in session or session['role'] != 'Kasir':
        return redirect(url_for('login'))

    user = session.get('username')

    # Ambil roti dari database
    roti_list = JenisRoti.query.all()

    return render_template("kasir/Hitung total & bayar.html", user=user, roti_list=roti_list)

@app.route('/simpan_transaksi', methods=['POST'])
def simpan_transaksi():
    if 'role' not in session or session['role'] != 'Kasir':
        return redirect(url_for('login'))

    data = request.json

    idUser = session.get('user_id')
    total = data['total']
    bayar = data['bayar']
    kembalian = data['kembalian']
    items = data['items']

    transaksi = TransaksiPenjualan(
        idUser=idUser,
        Total_Harga=total,
        Uang_Diterima=bayar,
        Kembalian=kembalian,
        Tanggal_Transaksi=waktu_wib()
    )
    db.session.add(transaksi)
    db.session.commit()

    idTransaksi = transaksi.idTransaksi_Penjualan

    for item in items:
        roti = JenisRoti.query.get(item['idRoti'])
        harga_asli = float(roti.Harga)
        subtotal = harga_asli * item['qty']

        detail = DetailTransaksi(
            id_transaksi=idTransaksi,
            id_roti=item['idRoti'],
            Jumlah=item['qty'],
            SubTotal=subtotal
        )
        db.session.add(detail)

        rk = db.session.execute(text("""
            SELECT rk.idRingkasan, rk.Stok_Aktual
            FROM ringkasan_baru rk
            JOIN produksi p ON rk.idProduksi = p.idProduksi
            WHERE p.idRoti = :idRoti
            ORDER BY rk.Tanggal DESC, rk.idRingkasan DESC
            LIMIT 1
        """), {"idRoti": item['idRoti']}).fetchone()

        stok_terakhir = rk[1] if rk else 0

        idProduksiTerbaru = db.session.execute(text("""
            SELECT idProduksi FROM produksi WHERE idRoti=:idRoti ORDER BY idProduksi DESC LIMIT 1
        """), {"idRoti": item['idRoti']}).scalar()

        new_rk = RingkasanBaru(
            idUser=idUser,
            idProduksi=idProduksiTerbaru,
            idTransaksi_Penjualan=idTransaksi,
            Total_Produksi=0,
            Total_Terjual=item['qty'],
            Stok_Aktual=stok_terakhir - item['qty'],
            Total_Uang_Masuk=subtotal
        )
        db.session.add(new_rk)

    db.session.commit()

    return {"status": "success"}


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
